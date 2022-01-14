import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color, Border, Side
from openpyxl.styles.fills import fills
import math

###################
##### 1. path #####
###################
base_path = '/Users/tony/Preprocessing/create_phase_xlsx/'

output_path = base_path + '/result_individual1/'
if not os.path.exists(output_path):
    os.makedirs(output_path)

def create_phase_xlsx():
    ###################################
    ##### 2. read phase list xlsx #####
    ###################################
    fname = os.path.join(base_path, 'Surgical_Phase_Gastrectomy.xlsx')
    load_wb = openpyxl.load_workbook(fname, data_only=True, read_only=True)
    load_sheet = load_wb["ops전달용영상 및 annotator list"]
    list_sheet = list(load_sheet)

    video_name = []     #hutom id  
    xlsx_sheet = []     #case
    timestamp = []
    individual1 = []    #annotator initials
    try:   
        if(load_sheet['C3'].value == 'hutomID'):
            get_cells = load_sheet['C4' : 'C81']
            for row in get_cells:
                for cell in row:
                    if cell.value is None:
                        continue
                    else: 
                        value = math.trunc(cell.value)
                        video_name.append(value)
        if(load_sheet['F3'].value == 'Timestamp'):
            get_cells = load_sheet['F4' : 'F81']
            for row in get_cells:
                for cell in row:
                    timestamp.append(cell.value)
        if(load_sheet['G3'].value == 'Case'):
            get_cells = load_sheet['G4' : 'G81']
            for row in get_cells:
                for cell in row:
                    xlsx_sheet.append(cell.value)
        if(load_sheet['K3'].value == 'Individual1'):
            get_cells = load_sheet['K4' : 'K81']
            for row in get_cells:
                for cell in row:
                    if cell.value is None:
                        continue
                    else:
                        individual1.append(cell.value)
        print('read all info on sheet')
    except:
        if(load_sheet['C3'].value != 'hutomID'):
            print('error in C3, check hutomID')
        if(load_sheet['F3'].value != 'Timestamp'):
            print('error in F3, check Timestamp')
        if(load_sheet['G3'].value != 'Case'):
            print('error in G3, check Case')
        if(load_sheet['K3'].value != 'Individual1'):
            print('error in K3, check Individual1')
    
    print(len(video_name), len(xlsx_sheet), len(timestamp), len(individual1))
    ##############################
    ##### 3. write xlsx file #####
    ##############################
    for pair_of_file in zip(video_name, individual1):          #video name별, individual1 xlsx file 생성
        write_wb = openpyxl.Workbook()
        filename = str(pair_of_file[0]) + '_Individual_' + str(pair_of_file[1]) + '.xlsx'

        ws = write_wb.active
        for pair_of_sheet in zip(xlsx_sheet, timestamp):        
            if str(pair_of_file[0]) in str(pair_of_sheet[0]):
                ws = write_wb.create_sheet()        #case별로 sheet 만들기
                ws.title = pair_of_sheet[0]
                write_ws = write_wb[pair_of_sheet[0]]
                
                time_ff = pair_of_sheet[1][9:11]        # frame
                time_ss = pair_of_sheet[1][6:8]         # sec
                time_mm = pair_of_sheet[1][3:5]         # min
                time_hh = pair_of_sheet[1][1:2]         # hour

                row_range = int(time_hh)*60*60*30 + int(time_mm)*60*30 + int(time_ss)*30 + int(time_ff)
                sec = 0
                min = 0
                hour = 0
                for i in range(row_range+1):
                    frame = 0
                    frame += i%30
                    if (i >= 30):
                        if(i%30 == 0):
                            sec += 1
                            if(sec%60 == 0):
                                min += 1
                                sec = 0
                                if(min%60 == 0 and sec%60 == 0):
                                    hour += 1
                                    min = 0
                    min_format = format(min, '02')  # sheet timestamp 양식 
                    sec_format = format(sec, '02')
                    frame_format = format(frame, '02')
                    time_frame = str(hour) + ':' + str(min_format) + ':' + str(sec_format) + ':' + str(frame_format)
                    write_ws.cell(row=i+3, column=1, value=time_frame)
                    for j in range(3):
                        ws.cell(row=i+3, column=j+1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            write_ws['A1'] = 'format (h:mm:ss:ff)'
            write_ws['B1'] = 'ARMES annotation'
            write_ws['B2'] = 'ARMES annotation ' + pair_of_file[1]
            write_ws['C2'] = 'Only ARMES annotation ' + pair_of_file[1]
            write_ws.merge_cells('B1:C1')       # cell merge

            # cell style
            ws.freeze_panes = "A3"     

            horizontal_border = Border(top=Side(style='thin'), 
                                        bottom=Side(style='thin')
                                        )
            vertical_border = Border(right=Side(style='thin'))

            #열 크기 지정
            col_widths = {"A":20, "B":30, "C":30}
            for col_name in col_widths:
                ws.column_dimensions[col_name].width = col_widths[col_name]


            ws['B1'].fill = PatternFill('solid', fgColor='FFFF00')      #cell coloring
            for i in range(3):
                ws.cell(row=2, column=i+1).fill = PatternFill('solid', fgColor='74DF00')


            for i in range(3):     # head's row, column align 
                for j in range(3):
                    ws.cell(row=i+1, column=j+1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        del write_wb['Sheet'] 
        write_wb.save(output_path + filename)
        print(filename)


if __name__ == '__main__':
    # main()
    create_phase_xlsx()